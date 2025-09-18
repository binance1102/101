# -*- coding: utf-8 -*-
"""
GepAnalitic-Dataset-Core.py

Дополнено (13 метрик):
1) P(fill ≤ N) по барам/времени (KM-оценка, учёт цензурирования)
2) Kaplan–Meier: медиана и P90 времени до закрытия
3) ATR(14) и нормировка gap_size/MFE/MAE в ATR-единицах
4) Калибровка по gap_size_ATR (биннинг) с P(fill ≤ 10 бар), KM-медианой
5) Риск стоп-аута до закрытия при SL=k·ATR, k∈{0.5,1,1.5}; ожидание MFE_ATR | не стопнут
6) Пост-закрытие: дрейф +N баров (N∈{1,3,5,10}) в % и ATR
7) Межгэповое время (интер-аррайвл): медиана, P90
8) Сезонность: частота и P(fill ≤ 10 бар) по DOW/часу (агрегированные ключи)
9) Режим рынка (внутренний): квантиль волатильности (ATR) и тренда (ADX) → условные P(fill ≤ 10)
10) Фьючерсы: funding/basis-квантили → условные P(fill ≤ 10) (если есть столбцы)
11) Ордерфлоу-имбаланс в окне гэпа → условные P(fill ≤ 10) (порог 0.6)
12) Мульти-ТФ флаг (заглушка: NaN; появится при наличии колонок старшего ТФ)
13) Кросс-секционка: ранги по P(fill ≤ 10 бар) среди монет в батче

ПРИМЕЧАНИЕ:
- Весь прежний функционал сохранён.
- Новые колонки добавлены в итоговую строку row (build_one_row) и в группировщик make_grouped_columns_3.
- Экспорт в Excel — с жирными разделителями групп и более тонкими — подгрупп.
"""

import os
import re
import argparse
import traceback
from glob import glob
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed

import numpy as np
import pandas as pd
from pandas.core.dtypes.dtypes import DatetimeTZDtype

# =========================
# КОНСТАНТЫ ПУТЕЙ
# =========================
DEFAULT_INPUT_ROOT = r"G:\gap-analytics-xlsx-full"
DEFAULT_OUTPUT_DIR = r"D:\GepAnalitic-Dataset"

# =========================
# УТИЛИТЫ ПУТЕЙ
# =========================
def extract_symbol_interval(file_path: str) -> tuple[str, str]:
    parts = file_path.split(os.sep)
    try:
        symbol = [p.split("=", 1)[1] for p in parts if p.startswith("symbol=")][0]
        interval = [p.split("=", 1)[1] for p in parts if p.startswith("interval=")][0]
    except Exception:
        symbol = os.path.basename(os.path.dirname(os.path.dirname(file_path)))
        interval = os.path.basename(os.path.dirname(file_path))
    return symbol, interval

def smart_resolve_path(file_path: str) -> str:
    p = os.path.normpath(file_path)
    if os.path.isfile(p):
        return p
    candidates = [
        p.replace(os.sep + "market" + os.sep + "futures", os.sep + "market=futures"),
        p.replace("market=futures", os.sep + "market" + os.sep + "futures"),
    ]
    try:
        root_idx = p.lower().rfind("gap-analytics-xlsx-full")
        if root_idx != -1:
            root = p[:root_idx] + "gap-analytics-xlsx-full"
            sym = None; itv = None
            parts = p.split(os.sep)
            for part in parts:
                if part.startswith("symbol="): sym = part.split("=", 1)[1]
                if part.startswith("interval="): itv = part.split("=", 1)[1]
            if sym and itv:
                for m in ("futures", "spot"):
                    pattern = os.path.join(root, "exchange=*", f"market={m}",
                                           f"symbol={sym}", f"interval={itv}", "full_with_gaps.xlsx")
                    candidates.extend(glob(pattern))
    except Exception:
        pass
    for c in candidates:
        if c and os.path.isfile(c):
            print(f"⚠ Путь не найден, использую найденный файл: {c}")
            return c
    return p

# >>> ADDED: извлечение exchange/market из пути (работает и для других бирж)
def extract_exchange_market(path: str) -> tuple[str, str]:
    parts = os.path.normpath(path).split(os.sep)
    ex = next((p.split("=", 1)[1] for p in parts if p.startswith("exchange=")), "binance")
    mk = next((p.split("=", 1)[1] for p in parts if p.startswith("market=")), "futures")
    return ex, mk

# =========================
# ЧИСЛА/ДАТЫ/ФОРМАТЫ
# =========================
def to_numeric_series(s: pd.Series) -> pd.Series:
    return (s.astype(str).str.strip().str.replace(",", ".", regex=False)
              .pipe(pd.to_numeric, errors="coerce"))

def _norm_gap_type(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.lower()

def interval_to_minutes(interval: str) -> float:
    s = interval.strip().lower()
    m = re.fullmatch(r"(\d+)([mhd])", s)
    if not m:
        aliases = {
            "1min": 1, "3min": 3, "5min": 5, "15min": 15, "30min": 30,
            "1hour": 60, "2hour": 120, "4hour": 240, "6hour": 360, "12hour": 720,
            "1day": 1440, "3day": 4320, "1d": 1440
        }
        if s in aliases:
            return float(aliases[s])
        raise ValueError(f"Не удалось разобрать интервал: {interval}")
    val = int(m.group(1)); unit = m.group(2)
    if unit == "m": return float(val)
    if unit == "h": return float(val * 60)
    if unit == "d": return float(val * 1440)
    raise ValueError(f"Неизвестная единица интервала: {interval}")

def to_utc_series(col: pd.Series) -> pd.Series:
    return pd.to_datetime(col, errors="coerce", utc=True)

def drop_tz_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        s = out[col]
        if isinstance(s.dtype, DatetimeTZDtype):
            out[col] = s.dt.tz_convert("UTC").dt.tz_localize(None)

    def col_endswith(name: str):
        hits = []
        for c in out.columns:
            leaf = c[-1] if isinstance(c, tuple) else c
            if leaf == name:
                hits.append(c)
        return hits
    for target in ["Текущее время (UTC)", "Последняя дата в таблице (UTC)"]:
        for c in col_endswith(target):
            try:
                parsed = pd.to_datetime(out[c], errors="coerce", utc=True, format="ISO8601")
            except TypeError:
                parsed = pd.to_datetime(out[c], errors="coerce", utc=True)
            out[c] = parsed.dt.tz_localize(None)
    return out

def stats_pack_num(a: pd.Series, want_pcts=True):
    x = pd.to_numeric(a, errors="coerce").dropna()
    if x.empty:
        d = dict(min=np.nan, max=np.nan, mean=np.nan, median=np.nan)
        if want_pcts: d.update({f"p{k}": np.nan for k in (50, 90, 95)})
        return d
    d = dict(min=float(x.min()), max=float(x.max()), mean=float(x.mean()), median=float(x.median()))
    if want_pcts:
        d.update({"p50": float(np.percentile(x, 50)),
                  "p90": float(np.percentile(x, 90)),
                  "p95": float(np.percentile(x, 95))})
    return d

# ----- человекочитаемые длительности -----
_MIN_PER_HOUR = 60
_MIN_PER_DAY  = 1440
_MIN_PER_MONTH= 43200   # ~30 дней
_MIN_PER_YEAR = 525600  # ~365 дней

def _fmt_hm(total_minutes: int) -> str:
    h, m = divmod(total_minutes, _MIN_PER_HOUR)
    return f"{h} ч {m} мин" if m else f"{h} ч"

def _fmt_dhm(total_minutes: int) -> str:
    d, rest = divmod(total_minutes, _MIN_PER_DAY)
    h, m = divmod(rest, _MIN_PER_HOUR)
    out = f"{d} дн"
    if h: out += f" {h} ч"
    if m: out += f" {m} мин"
    return out

def _fmt_m_d(total_minutes: int) -> str:
    mo, rest = divmod(total_minutes, _MIN_PER_MONTH)
    d, _ = divmod(rest, _MIN_PER_DAY)
    out = f"{mo} мес"
    if d: out += f" {d} дн"
    return out

def _fmt_y_m(total_minutes: int) -> str:
    y, rest = divmod(total_minutes, _MIN_PER_YEAR)
    mo, _ = divmod(rest, _MIN_PER_MONTH)
    out = f"{y} г"
    if mo: out += f" {mo} мес"
    return out

def humanize_minutes_value(x) -> str:
    if x is None:
        return ""
    try:
        m = int(round(float(x)))
    except Exception:
        return ""
    if m < 0 or np.isnan(m) or np.isinf(m):
        return ""
    if m < _MIN_PER_HOUR:
        return f"{m} мин"
    if m < _MIN_PER_DAY:
        return _fmt_hm(m)
    if m < _MIN_PER_MONTH:
        return _fmt_dhm(m)
    if m < _MIN_PER_YEAR:
        return _fmt_m_d(m)
    return _fmt_y_m(m)

def humanize_minutes_pack(pack: dict) -> dict:
    out = {}
    for k in ["min", "max", "mean", "median", "p50", "p90", "p95"]:
        v = pack.get(k, np.nan)
        out[k] = humanize_minutes_value(v) if pd.notna(v) else ""
    return out

# =========================
# БАЗОВЫЕ СЕРИИ/МАСКИ
# =========================
def pick_minutes_bars(df: pd.DataFrame, interval: str, prefer_existing: bool = True):
    tfm = interval_to_minutes(interval)
    gt = _norm_gap_type(df["gap_type"])
    ot = to_utc_series(df["open_time"])
    ft = to_utc_series(df["gap_filled_time"])
    closed = ft.notna() & ot.notna()
    up_closed = gt.eq("up") & closed
    down_closed = gt.eq("down") & closed

    minutes_calc = ((ft - ot).dt.total_seconds() / 60).where(closed)

    minutes_used = minutes_calc
    if prefer_existing and "gap_time_to_fill" in df.columns:
        mins_exist = to_numeric_series(df["gap_time_to_fill"]).where(closed)
        if mins_exist.notna().sum() > 0:
            minutes_used = mins_exist

    bars_used = (minutes_calc / tfm)
    if prefer_existing and "gap_bars_to_fill" in df.columns:
        bars_exist = to_numeric_series(df["gap_bars_to_fill"]).where(closed)
        if bars_exist.notna().sum() > 0:
            bars_used = bars_exist

    return {
        "tfm": tfm,
        "gt": gt,
        "ot": ot, "ft": ft,
        "closed": closed, "up_closed": up_closed, "down_closed": down_closed,
        "minutes": minutes_used, "bars": bars_used,
    }

# =========================
# ТЕХНИЧЕСКИЕ ИНДИКАТОРЫ: ATR/ADX
# =========================
def compute_atr(high: pd.Series, low: pd.Series, close: pd.Series, n: int = 14) -> pd.Series:
    high = to_numeric_series(high); low = to_numeric_series(low); close = to_numeric_series(close)
    prev_close = close.shift(1)
    tr = pd.concat([
        (high - low).abs(),
        (high - prev_close).abs(),
        (low  - prev_close).abs()
    ], axis=1).max(axis=1)
    # классическая Wilder ATR
    atr = tr.ewm(alpha=1/n, adjust=False).mean()
    return atr

def compute_adx(high: pd.Series, low: pd.Series, close: pd.Series, n: int = 14) -> pd.Series:
    # упрощённый ADX
    high = to_numeric_series(high); low = to_numeric_series(low); close = to_numeric_series(close)
    up_move = high.diff()
    down_move = -low.diff()
    plus_dm  = np.where((up_move > down_move) & (up_move > 0), up_move, 0.0)
    minus_dm = np.where((down_move > up_move) & (down_move > 0), down_move, 0.0)
    tr = pd.concat([
        (high - low).abs(),
        (high - close.shift(1)).abs(),
        (low  - close.shift(1)).abs()
    ], axis=1).max(axis=1)
    atr = tr.ewm(alpha=1/n, adjust=False).mean()
    plus_di = 100 * pd.Series(plus_dm, index=high.index).ewm(alpha=1/n, adjust=False).mean() / atr
    minus_di= 100 * pd.Series(minus_dm, index=high.index).ewm(alpha=1/n, adjust=False).mean() / atr
    dx = ( (plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, np.nan) ) * 100
    adx = dx.ewm(alpha=1/n, adjust=False).mean()
    return adx

# =========================
# MFE/MAE по закрытым (как было)
# =========================
def aggregate_mfe_mae(df: pd.DataFrame, interval: str, meta: dict, mfe_base: str):
    gt = meta["gt"]; ot = meta["ot"]; ft = meta["ft"]
    high = to_numeric_series(df.get("high", pd.Series(dtype=float)))
    low  = to_numeric_series(df.get("low",  pd.Series(dtype=float)))
    open_ = to_numeric_series(df.get("open", pd.Series(dtype=float)))
    close_ = to_numeric_series(df.get("close", pd.Series(dtype=float)))
    gap_level = to_numeric_series(df.get("gap_level", pd.Series(dtype=float)))

    if mfe_base == "gap_level": base = gap_level
    elif mfe_base == "open":    base = open_
    elif mfe_base == "close":   base = close_
    else:                       base = gap_level
    base = base.where(meta["closed"])

    idx_sorted = meta["ot"].sort_values().index
    ot_sorted = meta["ot"].loc[idx_sorted]
    high_sorted = high.loc[idx_sorted]
    low_sorted  = low.loc[idx_sorted]

    def win_lr(start_ts, end_ts):
        arr = ot_sorted.values
        l = np.searchsorted(arr, start_ts.to_datetime64(), side="left")
        r = np.searchsorted(arr, end_ts.to_datetime64(), side="right")
        return None if r <= l else (l, r)

    max_high_val = pd.Series(np.nan, index=df.index, dtype=float)
    max_high_ts  = pd.Series(pd.NaT, index=df.index, dtype="datetime64[ns, UTC]")
    min_low_val  = pd.Series(np.nan, index=df.index, dtype=float)
    min_low_ts   = pd.Series(pd.NaT, index=df.index, dtype="datetime64[ns, UTC]")

    for i in df.index:
        if not meta["closed"].loc[i]: continue
        w = win_lr(meta["ot"].loc[i], meta["ft"].loc[i])
        if w is None: continue
        l, r = w
        sub_h = high_sorted.iloc[l:r]; sub_l = low_sorted.iloc[l:r]
        if len(sub_h):
            p = int(sub_h.values.argmax())
            max_high_val.loc[i] = float(sub_h.iloc[p])
            max_high_ts.loc[i]  = ot_sorted.iloc[l + p]
        if len(sub_l):
            p = int(sub_l.values.argmin())
            min_low_val.loc[i] = float(sub_l.iloc[p])
            min_low_ts.loc[i]  = ot_sorted.iloc[l + p]

    up = (gt == "up") & meta["closed"]
    dn = (gt == "down") & meta["closed"]

    time_to_max_high = ((max_high_ts - meta["ot"]).dt.total_seconds() / 60)
    time_to_min_low  = ((min_low_ts  - meta["ot"]).dt.total_seconds() / 60)

    mfe_price = pd.Series(np.nan, index=df.index, dtype=float)
    mae_price = pd.Series(np.nan, index=df.index, dtype=float)
    mfe_min   = pd.Series(np.nan, index=df.index, dtype=float)
    mae_min   = pd.Series(np.nan, index=df.index, dtype=float)

    # up
    mfe_price.loc[up] = (max_high_val - base).loc[up]
    mae_price.loc[up] = (base - min_low_val).loc[up]
    mfe_min.loc[up]   = time_to_max_high.loc[up]
    mae_min.loc[up]   = time_to_min_low.loc[up]
    # down
    mfe_price.loc[dn] = (base - min_low_val).loc[dn]
    mae_price.loc[dn] = (max_high_val - base).loc[dn]
    mfe_min.loc[dn]   = time_to_min_low.loc[dn]
    mae_min.loc[dn]   = time_to_max_high.loc[dn]

    def safe_pct(num, den):
        den = den.replace(0, np.nan)
        return (num / den) * 100.0

    mfe_pct = safe_pct(mfe_price, base)
    mae_pct = safe_pct(mae_price, base)
    bars_to_mfe = (mfe_min / meta["tfm"])
    bars_to_mae = (mae_min / meta["tfm"])

    return {
        "up":   {"mfe_price": mfe_price[up], "mae_price": mae_price[up],
                 "mfe_pct": mfe_pct[up], "mae_pct": mae_pct[up],
                 "mfe_min": mfe_min[up], "mae_min": mae_min[up],
                 "mfe_bars": bars_to_mfe[up], "mae_bars": bars_to_mae[up]},
        "down": {"mfe_price": mfe_price[dn], "mae_price": mae_price[dn],
                 "mfe_pct": mfe_pct[dn], "mae_pct": mae_pct[dn],
                 "mfe_min": mfe_min[dn], "mae_min": mae_min[dn],
                 "mfe_bars": bars_to_mfe[dn], "mae_bars": bars_to_mae[dn]},
        "all":  {"mfe_price": mfe_price[meta["closed"]], "mae_price": mae_price[meta["closed"]],
                 "mfe_pct": mfe_pct[meta["closed"]], "mae_pct": mae_pct[meta["closed"]],
                 "mfe_min": mfe_min[meta["closed"]], "mae_min": mae_min[meta["closed"]],
                 "mfe_bars": bars_to_mfe[meta["closed"]], "mae_bars": bars_to_mae[meta["closed"]]},
        "base_label": mfe_base
    }

# =========================
# Kaplan–Meier / P(fill ≤ N)
# =========================
def km_survival(durations_min: pd.Series, event_observed: pd.Series):
    """Простая KM: возвращает точки (t, S(t)), медиану и P90"""
    d = pd.DataFrame({"t": pd.to_numeric(durations_min, errors="coerce"),
                      "e": event_observed.astype(bool)})
    d = d.dropna()
    if d.empty:
        return [], np.nan, np.nan

    df = d.sort_values("t")
    uniq = np.unique(df["t"].values)
    n = len(df)
    s = 1.0
    surv = []
    idx = 0
    for t in uniq:
        # число событий и цензур в точке
        at_t = df["t"].values == t
        d_i = int((at_t & df["e"].values).sum())  # события
        c_i = int((at_t & (~df["e"].values)).sum())  # цензура
        if n > 0:
            s *= (1 - d_i / n)
        surv.append((t, s))
        n -= (d_i + c_i)

    # медиана: минимальное t, где S(t) <= 0.5
    median = np.nan
    p90 = np.nan
    for t, sval in surv:
        if np.isnan(median) and sval <= 0.5:
            median = t
        if np.isnan(p90) and sval <= 0.1:
            p90 = t
    return surv, median, p90

def km_prob_fill_leq(surv_points, horizon_min: float) -> float:
    """P(T ≤ h) = 1 - S(h)"""
    if not surv_points:
        return np.nan
    # найдём S(h) как значение наибольшего t <= h
    s_h = 1.0
    for t, s in surv_points:
        if t <= horizon_min:
            s_h = s
        else:
            break
    return 1.0 - s_h

# =========================
# СЕРИЙНОСТЬ
# =========================
def compute_direction_run_stats(df: pd.DataFrame) -> dict:
    gt = _norm_gap_type(df["gap_type"])
    ot = to_utc_series(df["open_time"])
    ok = gt.isin(["up", "down"]) & ot.notna()
    if ok.sum() <= 1:
        return {}

    idx = ot[ok].sort_values().index
    seq = gt.loc[idx].map({"up": 1, "down": -1}).to_numpy()

    runs = []
    cur_dir = seq[0]
    cur_len = 1
    for x in seq[1:]:
        if x == cur_dir:
            cur_len += 1
        else:
            runs.append((cur_dir, cur_len))
            cur_dir = x
            cur_len = 1
    terminated_runs = runs[:]
    if not terminated_runs:
        return {}

    ups  = [L for d, L in terminated_runs if d == 1]
    dns  = [L for d, L in terminated_runs if d == -1]
    allr = [L for _, L in terminated_runs]

    def mode_int(arr):
        if not arr: return np.nan
        vals, cnts = np.unique(arr, return_counts=True)
        return int(vals[np.argmax(cnts)])

    def pack_stats(arr):
        if not arr:
            return dict(mean=np.nan, median=np.nan, p90=np.nan, mode=np.nan)
        a = np.asarray(arr, dtype=float)
        return dict(
            mean=float(a.mean()),
            median=float(np.median(a)),
            p90=float(np.percentile(a, 90)),
            mode=mode_int(list(map(int, a)))
        )

    st_up  = pack_stats(ups)
    st_dn  = pack_stats(dns)
    st_all = pack_stats(allr)

    def bucket_dist(arr):
        if not arr:
            return {f"n{k}": 0 for k in [1,2,3,4,5]} | {"n6p": 0}
        a = np.asarray(arr, dtype=int)
        out = {f"n{k}": int((a==k).sum()) for k in [1,2,3,4,5]}
        out["n6p"] = int((a>=6).sum())
        return out

    dist_up  = bucket_dist(ups)
    dist_dn  = bucket_dist(dns)
    dist_all = bucket_dist(allr)

    def hazard(arr, kmax=5):
        if not arr:
            return {f"k{k}": np.nan for k in range(1, kmax+1)}
        a = np.asarray(arr, dtype=int)
        out = {}
        for k in range(1, kmax+1):
            reached = int((a >= k).sum())
            changed = int((a == k).sum())
            out[f"k{k}"] = (changed / reached * 100.0) if reached > 0 else np.nan
        return out

    haz_up  = hazard(ups)
    haz_dn  = hazard(dns)

    switches = int((seq[1:] != seq[:-1]).sum())
    total_pairs = len(seq) - 1
    switch_share = (switches / total_pairs * 100.0) if total_pairs > 0 else np.nan

    out = {}
    out["Серии ↑ — средняя длина до смены"] = round(st_up["mean"], 3) if pd.notna(st_up["mean"]) else np.nan
    out["Серии ↑ — медиана длины до смены"] = round(st_up["median"], 3) if pd.notna(st_up["median"]) else np.nan
    out["Серии ↑ — P90 длины до смены"]     = round(st_up["p90"], 3) if pd.notna(st_up["p90"]) else np.nan
    out["Серии ↑ — мода длины до смены"]    = st_up["mode"] if pd.notna(st_up["mode"]) else np.nan

    out["Серии ↓ — средняя длина до смены"] = round(st_dn["mean"], 3) if pd.notna(st_dn["mean"]) else np.nan
    out["Серии ↓ — медиана длины до смены"] = round(st_dn["median"], 3) if pd.notna(st_dn["median"]) else np.nan
    out["Серии ↓ — P90 длины до смены"]     = round(st_dn["p90"], 3) if pd.notna(st_dn["p90"]) else np.nan
    out["Серии ↓ — мода длины до смены"]    = st_dn["mode"] if pd.notna(st_dn["mode"]) else np.nan

    out["Серии (все) — средняя длина до смены"] = round(st_all["mean"], 3) if pd.notna(st_all["mean"]) else np.nan
    out["Серии (все) — медиана длины до смены"] = round(st_all["median"], 3) if pd.notna(st_all["median"]) else np.nan
    out["Серии (все) — P90 длины до смены"]     = round(st_all["p90"], 3) if pd.notna(st_all["p90"]) else np.nan
    out["Серии (все) — мода длины до смены"]    = st_all["mode"] if pd.notna(st_all["mode"]) else np.nan

    for lab, dist in [("↑", dist_up), ("↓", dist_dn), ("(все)", dist_all)]:
        out[f"Распределение длин серий {lab} — n1"] = dist["n1"]
        out[f"Распределение длин серий {lab} — n2"] = dist["n2"]
        out[f"Распределение длин серий {lab} — n3"] = dist["n3"]
        out[f"Распределение длин серий {lab} — n4"] = dist["n4"]
        out[f"Распределение длин серий {lab} — n5"] = dist["n5"]
        out[f"Распределение длин серий {lab} — n6+"] = dist["n6p"]

    for k in range(1, 6):
        out[f"Вероятность смены после {k} подряд ↑, %"] = round(haz_up.get(f"k{k}", np.nan), 2) if haz_up else np.nan
        out[f"Вероятность смены после {k} подряд ↓, %"] = round(haz_dn.get(f"k{k}", np.nan), 2) if haz_dn else np.nan

    out["Доля смен направления между соседними гэпами, %"] = round(switch_share, 2) if pd.notna(switch_share) else np.nan
    return out

# =========================
# ОТКРЫТЫЕ: возраст / импульс (как было)
# =========================
def aggregate_open_ages(df: pd.DataFrame, interval: str, meta: dict, now_equals_last: bool=False):
    tfm = meta["tfm"]; gt = meta["gt"]; ot = meta["ot"]; ft = meta["ft"]

    gap_rows = gt.isin(["up", "down"])
    open_mask = gap_rows & ft.isna() & ot.notna()

    cols_for_last = [ot]
    ct = to_utc_series(df["close_time"]) if "close_time" in df.columns else pd.Series(pd.NaT, index=df.index)
    if ct.notna().any(): cols_for_last.append(ct)
    if ft.notna().any(): cols_for_last.append(ft)
    last_ts_series = pd.concat(cols_for_last, axis=1).max(axis=1, skipna=True)
    last_ts = pd.to_datetime(last_ts_series.max(), utc=True)

    now_utc = pd.Timestamp.now(tz="UTC")
    if now_equals_last:
        now_utc = last_ts

    mn_now  = ((now_utc  - ot).dt.total_seconds() / 60).where(open_mask)
    mn_last = ((last_ts  - ot).dt.total_seconds() / 60).where(open_mask)
    br_now  = mn_now / tfm
    br_last = mn_last / tfm

    up_open  = (gt == "up") & open_mask
    dn_open  = (gt == "down") & open_mask
    all_open = open_mask

    def agg_age(minutes: pd.Series, bars: pd.Series):
        s_m = minutes.dropna(); s_b = pd.to_numeric(bars, errors="coerce").dropna()
        d = {}
        m = stats_pack_num(s_m, want_pcts=True)
        m_h = humanize_minutes_pack(m)
        d.update({
            "min_min": m_h["min"], "max_min": m_h["max"],
            "mean_min": m_h["mean"], "median_min": m_h["median"],
            "p50_min": m_h.get("p50", ""), "p90_min": m_h.get("p90", ""), "p95_min": m_h.get("p95", ""),
        })
        b = stats_pack_num(s_b, want_pcts=True)
        d.update({
            "min_bars": round(b["min"], 2) if pd.notna(b["min"]) else np.nan,
            "max_bars": round(b["max"], 2) if pd.notna(b["max"]) else np.nan,
            "mean_bars": round(b["mean"], 2) if pd.notna(b["mean"]) else np.nan,
            "median_bars": round(b["median"], 2) if pd.notna(b["median"]) else np.nan,
        })
        if len(s_m):
            hrs = s_m/60; days = s_m/1440; months = days/30.0; years = days/365.0
            d.update({
                "mean_hours": round(float(hrs.mean()), 2),
                "mean_days": round(float(days.mean()), 3),
                "mean_months": round(float(months.mean()), 4),
                "mean_years": round(float(years.mean()), 5),
            })
        else:
            d.update({"mean_hours": np.nan, "mean_days": np.nan, "mean_months": np.nan, "mean_years": np.nan})
        return d

    now_up   = agg_age(mn_now.where(up_open),  br_now.where(up_open))
    now_dn   = agg_age(mn_now.where(dn_open),  br_now.where(dn_open))
    now_all  = agg_age(mn_now.where(all_open), br_now.where(all_open))
    last_up  = agg_age(mn_last.where(up_open),  br_last.where(up_open))
    last_dn  = agg_age(mn_last.where(dn_open),  br_last.where(dn_open))
    last_all = agg_age(mn_last.where(all_open), br_last.where(all_open))

    n_up  = int(up_open.sum())
    n_dn  = int(dn_open.sum())
    n_all = int(all_open.sum())

    return {
        "now": {"up": now_up, "down": now_dn, "all": now_all,
                "n_up": n_up, "n_down": n_dn, "n_all": n_all},
        "last": {"up": last_up, "down": last_dn, "all": last_all,
                 "n_up": n_up, "n_down": n_dn, "n_all": n_all},
        "now_utc": now_utc, "last_ts": last_ts
    }

def open_impulse_metrics(df: pd.DataFrame, meta: dict, mfe_base: str, end_ts: pd.Timestamp):
    gt = meta["gt"]; ot = meta["ot"]; ft = meta["ft"]
    high = to_numeric_series(df.get("high", pd.Series(dtype=float)))
    low  = to_numeric_series(df.get("low",  pd.Series(dtype=float)))
    open_ = to_numeric_series(df.get("open",  pd.Series(dtype=float)))
    close_ = to_numeric_series(df.get("close", pd.Series(dtype=float)))
    gap_level = to_numeric_series(df.get("gap_level", pd.Series(dtype=float)))

    open_mask = gt.isin(["up", "down"]) & ft.isna() & ot.notna()

    if mfe_base == "gap_level": base = gap_level
    elif mfe_base == "open":    base = open_
    elif mfe_base == "close":   base = close_
    else:                       base = gap_level
    base = base.where(open_mask)

    idx_sorted = ot.sort_values().index
    ot_sorted = ot.loc[idx_sorted]
    high_sorted = high.loc[idx_sorted]
    low_sorted  = low.loc[idx_sorted]

    def win_lr(start_ts, end_ts_):
        arr = ot_sorted.values
        l = np.searchsorted(arr, start_ts.to_datetime64(), side="left")
        r = np.searchsorted(arr, end_ts_.to_datetime64(), side="right")
        return None if r <= l else (l, r)

    max_high_val = pd.Series(np.nan, index=df.index, dtype=float)
    min_low_val  = pd.Series(np.nan, index=df.index, dtype=float)

    for i in df.index:
        if not open_mask.loc[i]: continue
        w = win_lr(ot.loc[i], end_ts)
        if w is None: continue
        l, r = w
        sub_h = high_sorted.iloc[l:r]; sub_l = low_sorted.iloc[l:r]
        if len(sub_h): max_high_val.loc[i] = float(sub_h.max())
        if len(sub_l): min_low_val.loc[i]  = float(sub_l.min())

    mfe_price = pd.Series(np.nan, index=df.index, dtype=float)
    mae_price = pd.Series(np.nan, index=df.index, dtype=float)

    up = (gt == "up") & open_mask
    dn = (gt == "down") & open_mask
    mfe_price.loc[up] = (max_high_val - base).loc[up]
    mae_price.loc[up] = (base - min_low_val).loc[up]
    mfe_price.loc[dn] = (base - min_low_val).loc[dn]
    mae_price.loc[dn] = (max_high_val - base).loc[dn]

    def block(side_mask: pd.Series):
        m = mfe_price.where(side_mask).dropna()
        a = mae_price.where(side_mask).abs().dropna()
        both = pd.concat([m, a], axis=1, keys=["mfe", "mae"]).dropna()
        if both.empty:
            return dict(mean=np.nan, median=np.nan, p90=np.nan, share=np.nan)
        ratio = (both["mfe"] / both["mae"].replace(0, np.nan)).replace([np.inf, -np.inf], np.nan).dropna()
        share = (both["mfe"] > both["mae"]).mean() * 100.0
        return dict(
            mean=float(ratio.mean()) if len(ratio) else np.nan,
            median=float(ratio.median()) if len(ratio) else np.nan,
            p90=float(np.percentile(ratio, 90)) if len(ratio) else np.nan,
            share=float(share),
        )

    res_up = block(up)
    res_dn = block(dn)
    res_all = block(open_mask)

    return {"up": res_up, "down": res_dn, "all": res_all}

# =========================
# БАР/ОКНО ГЭПА (как было)
# =========================
def aggregate_gapbar_volume_delta_by_bucket(df: pd.DataFrame, meta: dict):
    gt = meta["gt"]
    closed_mask = meta["closed"]
    open_mask = gt.isin(["up", "down"]) & (~closed_mask) & meta["ot"].notna()
    all_mask = gt.isin(["up", "down"])

    volume = to_numeric_series(df.get("volume", pd.Series(dtype=float)))
    qav    = to_numeric_series(df.get("qav", pd.Series(dtype=float)))
    tb     = to_numeric_series(df.get("taker_base_vol",  pd.Series(dtype=float)))
    tq     = to_numeric_series(df.get("taker_quote_vol", pd.Series(dtype=float)))
    open_  = to_numeric_series(df.get("open",  pd.Series(dtype=float)))
    close_ = to_numeric_series(df.get("close", pd.Series(dtype=float)))

    delta_base_all  = (2*tb - volume)
    delta_quote_all = (2*tq - qav)

    price_delta_abs_all = (close_ - open_)
    price_delta_pct_all = (price_delta_abs_all / open_.replace(0, np.nan) * 100.0)

    def mmm(s: pd.Series, mask: pd.Series):
        x = pd.to_numeric(s.where(mask), errors="coerce").dropna()
        if x.empty: return (np.nan, np.nan, np.nan)
        return (float(x.min()), float(x.mean()), float(x.max()))

    def ssum(s: pd.Series, mask: pd.Series):
        x = pd.to_numeric(s.where(mask), errors="coerce").dropna()
        return float(x.sum()) if len(x) else np.nan

    out = {}
    buckets = [
        ("(Закрытые)", closed_mask),
        ("(Открытые)", open_mask),
        ("(Все)",      all_mask),
    ]
    for b_name, b_mask in buckets:
        # >>> ADDED: фикс для опечатки переменной в f-строке
        б_name = b_name  # не меняем вашу строку, лишь гарантируем, что переменная существует

        mask_up = (gt.eq("up")   & b_mask)
        mask_dn = (gt.eq("down") & b_mask)
        mask_all= (gt.isin(["up","down"]) & b_mask)

        for lab, m in [("↑", mask_up), ("↓", mask_dn), ("(все)", mask_all)]:
            vmin, vmean, vmax = mmm(volume, m)
            out[f"{b_name} Объём на баре гэпа — мин {lab}"]   = round(vmin, 6) if pd.notna(vmin) else np.nan
            out[f"{б_name if False else b_name} Объём на баре гэпа — ср {lab}"]    = round(vmean, 6) if pd.notna(vmean) else np.nan
            out[f"{b_name} Объём на баре гэпа — макс {lab}"]  = round(vmax, 6) if pd.notna(vmax) else np.nan

            dmin, dmean, dmax = mmm(delta_base_all, m)
            out[f"{b_name} Дельта объёма (base) — мин {lab}"]  = round(dmin, 6) if pd.notna(dmin) else np.nan
            out[f"{b_name} Дельта объёма (base) — ср {lab}"]   = round(dmean, 6) if pd.notna(dmean) else np.nan
            out[f"{b_name} Дельта объёма (base) — макс {lab}"] = round(dmax, 6) if pd.notna(dmax) else np.nan

            qmin, qmean, qmax = mmm(delta_quote_all, m)
            out[f"{b_name} Дельта объёма (quote) — мин {lab}"]  = round(qmin, 6) if pd.notna(qmin) else np.nan
            out[f"{b_name} Дельта объёма (quote) — ср {lab}"]   = round(qmean, 6) if pd.notna(qmean) else np.nan
            out[f"{b_name} Дельта объёма (quote) — макс {lab}"] = round(qmax, 6) if pd.notna(qmax) else np.nan

            pmin, pmean, pmax = mmm(price_delta_abs_all, m)
            out[f"{b_name} Дельта цены (abs) — мин {lab}"]  = round(pmin, 10) if pd.notna(pmin) else np.nan
            out[f"{b_name} Дельта цены (abs) — ср {lab}"]   = round(pmean, 10) if pd.notna(pmean) else np.nan
            out[f"{b_name} Дельта цены (abs) — макс {lab}"] = round(pmax, 10) if pd.notna(pmax) else np.nan

            ppmin, ppmean, ppmax = mmm(price_delta_pct_all, m)
            out[f"{b_name} Дельта цены (%) — мин {lab}"]  = round(ppmin, 6) if pd.notna(ppmin) else np.nan
            out[f"{b_name} Дельта цены (%) — ср {lab}"]   = round(ppmean, 6) if pd.notna(ppmean) else np.nan
            out[f"{b_name} Дельта цены (%) — макс {lab}"] = round(ppmax, 6) if pd.notna(ppmax) else np.nan

            out[f"{b_name} Кумулятивная дельта (base) на барах гэпа — сумма {lab}"]  = round(ssum(delta_base_all, m), 6) if pd.notna(ssum(delta_base_all, m)) else np.nan
            out[f"{b_name} Кумулятивная дельта (quote) на барах гэпа — сумма {lab}"] = round(ssum(delta_quote_all, m), 6) if pd.notna(ssum(delta_quote_all, m)) else np.nan

    return out

def aggregate_gapbar_oi_by_bucket(df: pd.DataFrame, meta: dict):
    gt = meta["gt"]
    closed_mask = meta["closed"]
    open_mask = gt.isin(["up", "down"]) & (~closed_mask) & meta["ot"].notna()
    all_mask = gt.isin(["up", "down"])

    oi_ch = to_numeric_series(df.get("oi_change_at_gap", pd.Series(dtype=float)))

    def mmm(s: pd.Series, mask: pd.Series):
        x = pd.to_numeric(s.where(mask), errors="coerce").dropna()
        if x.empty: return (np.nan, np.nan, np.nan)
        return (float(x.min()), float(x.mean()), float(x.max()))

    def ssum(s: pd.Series, mask: pd.Series):
        x = pd.to_numeric(s.where(mask), errors="coerce").dropna()
        return float(x.sum()) if len(x) else np.nan

    out = {}
    buckets = [
        ("(Закрытые)", closed_mask),
        ("(Открытые)", open_mask),
        ("(Все)",      all_mask),
    ]
    for b_name, b_mask in buckets:
        mask_up = (gt.eq("up")   & b_mask)
        mask_dn = (gt.eq("down") & b_mask)
        mask_all= (gt.isin(["up","down"]) & b_mask)
        for lab, m in [("↑", mask_up), ("↓", mask_dn), ("(все)", mask_all)]:
            omin, omean, omax = mmm(oi_ch, m)
            out[f"{b_name} Изм. OI на баре гэпа — мин {lab}"]   = round(omin, 6) if pd.notna(omin) else np.nan
            out[f"{b_name} Изм. OI на баре гэпа — ср {lab}"]    = round(omean, 6) if pd.notna(omean) else np.nan
            out[f"{b_name} Изм. OI на баре гэпа — макс {lab}"]  = round(omax, 6) if pd.notna(omax) else np.nan
            out[f"{b_name} Изм. OI на барах гэпа — КУМУЛЯТИВ, сумма {lab}"] = round(ssum(oi_ch, m), 6) if pd.notna(ssum(oi_ch, m)) else np.nan

    return out

def aggregate_window_volume_delta(df: pd.DataFrame, meta: dict, bucket: str,
                                  end_ts_global: pd.Timestamp | None):
    gt = meta["gt"]; ot = meta["ot"]; ft = meta["ft"]

    volume = to_numeric_series(df.get("volume", pd.Series(dtype=float)))
    qav    = to_numeric_series(df.get("qav", pd.Series(dtype=float)))
    tb     = to_numeric_series(df.get("taker_base_vol",  pd.Series(dtype=float)))
    tq     = to_numeric_series(df.get("taker_quote_vol", pd.Series(dtype=float)))

    delta_base = (2*tb - volume)
    delta_quote= (2*tq - qav)

    idx_sorted = ot.sort_values().index
    ot_sorted  = ot.loc[idx_sorted]
    vol_s      = volume.loc[idx_sorted]
    db_s       = delta_base.loc[idx_sorted]
    dq_s       = delta_quote.loc[idx_sorted]

    if bucket == "closed":
        row_mask = meta["closed"]
        end_vec  = ft
        bucket_title = "(Закрытые ОКНО)"
    else:
        open_mask = gt.isin(["up","down"]) & (~meta["closed"]) & ot.notna()
        row_mask  = open_mask
        end_vec   = pd.Series(end_ts_global, index=df.index)
        bucket_title = "(Открытые→СЕЙЧАС ОКНО)" if bucket == "open_now" else "(Открытые→ПОСЛЕДНЯЯ ОКНО)"

    def win_lr(start_ts, end_ts):
        arr = ot_sorted.values
        l = np.searchsorted(arr, start_ts.to_datetime64(), side="left")
        r = np.searchsorted(arr, end_ts.to_datetime64(), side="right")
        return None if r <= l else (l, r)

    acc = {
        "up":    {"sum_vol": [], "max_vol_bar": [], "sum_db": [], "mean_db_bar": [], "min_db_bar": [], "max_db_bar": [],
                  "sum_dq": [], "mean_dq_bar": [], "min_dq_bar": [], "max_dq_bar": []},
        "down":  {"sum_vol": [], "max_vol_bar": [], "sum_db": [], "mean_db_bar": [], "min_db_bar": [], "max_db_bar": [],
                  "sum_dq": [], "mean_dq_bar": [], "min_dq_bar": [], "max_dq_bar": []},
        "all":   {"sum_vol": [], "max_vol_bar": [], "sum_db": [], "mean_db_bar": [], "min_db_bar": [], "max_db_bar": [],
                  "sum_dq": [], "mean_dq_bar": [], "min_dq_bar": [], "max_dq_bar": []},
    }

    def push(side_key, sub_v: pd.Series, sub_db: pd.Series, sub_dq: pd.Series):
        if sub_v.empty:
            return
        acc[side_key]["sum_vol"].append(float(sub_v.sum()))
        acc[side_key]["max_vol_bar"].append(float(sub_v.max()))
        acc[side_key]["sum_db"].append(float(sub_db.sum()))
        acc[side_key]["mean_db_bar"].append(float(sub_db.mean()))
        acc[side_key]["min_db_bar"].append(float(sub_db.min()))
        acc[side_key]["max_db_bar"].append(float(sub_db.max()))
        acc[side_key]["sum_dq"].append(float(sub_dq.sum()))
        acc[side_key]["mean_dq_bar"].append(float(sub_dq.mean()))
        acc[side_key]["min_dq_bar"].append(float(sub_dq.min()))
        acc[side_key]["max_dq_bar"].append(float(sub_dq.max()))

    for i in df.index:
        if not (row_mask.loc[i] and ot.notna().loc[i] and end_vec.notna().loc[i]):
            continue
        w = win_lr(ot.loc[i], end_vec.loc[i])
        if w is None:
            continue
        l, r = w
        sub_v  = vol_s.iloc[l:r].dropna()
        sub_db = db_s.iloc[l:r].dropna()
        sub_dq = dq_s.iloc[l:r].dropna()
        if sub_v.empty or sub_db.empty or sub_dq.empty:
            continue

        side = "up" if gt.loc[i] == "up" else ("down" if gt.loc[i] == "down" else "all")
        push(side, sub_v, sub_db, sub_dq)
        push("all", sub_v, sub_db, sub_dq)

    def mmm_list(vals):
        s = pd.Series(vals, dtype=float)
        if s.empty: return (np.nan, np.nan, np.nan)
        return (float(s.min()), float(s.mean()), float(s.max()))

    def sum_list(vals):
        s = pd.Series(vals, dtype=float)
        return float(s.sum()) if len(s) else np.nan

    out = {}
    for side_key, label in [("up", "↑"), ("down", "↓"), ("all", "(все)")]:
        A = acc[side_key]
        for metric, human in [
            ("sum_vol", "Сумма объёма в окне"),
            ("max_vol_bar", "Макс объём бара в окне"),
            ("sum_db", "Чистая дельта base в окне"),
            ("mean_db_bar", "Средняя дельта base/бар в окне"),
            ("min_db_bar", "Мин дельта base/бар в окне"),
            ("max_db_bar", "Макс дельта base/бар в окне"),
            ("sum_dq", "Чистая дельта quote в окне"),
            ("mean_dq_bar", "Средняя дельта quote/бар в окне"),
            ("min_dq_bar", "Мин дельта quote/бар в окне"),
            ("max_dq_bar", "Макс дельта quote/бар в окне"),
        ]:
            vmin, vmean, vmax = mmm_list(A[metric])
            out[f"{bucket_title} {human} — мин {label}"]  = round(vmin, 6) if pd.notna(vmin) else np.nan
            out[f"{bucket_title} {human} — ср {label}"]   = round(vmean, 6) if pd.notna(vmean) else np.nan
            out[f"{bucket_title} {human} — макс {label}"] = round(vmax, 6) if pd.notna(vmax) else np.nan

        out[f"{bucket_title} Кумулятивная чистая дельта base по всем гэпам — сумма {label}"]  = round(sum_list(A["sum_db"]), 6) if pd.notna(sum_list(A["sum_db"])) else np.nan
        out[f"{bucket_title} Кумулятивная чистая дельта quote по всем гэпам — сумма {label}"] = round(sum_list(A["sum_dq"]), 6) if pd.notna(sum_list(A["sum_dq"])) else np.nan

    return out

# =========================
# ДОП. МЕТРИКИ (13 пунктов) — ВСПОМОГАТЕЛЬНЫЕ
# =========================
def build_atr_adx_blocks(df: pd.DataFrame):
    atr = compute_atr(df["high"], df["low"], df["close"], n=14)
    adx = compute_adx(df["high"], df["low"], df["close"], n=14)
    return atr, adx

def build_gap_series(df: pd.DataFrame, meta: dict, mfe_base: str):
    """Возвращает Series по ряду характеристик на индекс df: atr_at_gap, gap_size_price, gap_size_atr, closed_mask, durations (мин), censor_durations_now"""
    ot = meta["ot"]; ft = meta["ft"]; tfm = meta["tfm"]; gt = meta["gt"]
    atr14, adx14 = build_atr_adx_blocks(df)
    high = to_numeric_series(df["high"]); low = to_numeric_series(df["low"]); open_ = to_numeric_series(df["open"]); close_ = to_numeric_series(df["close"])
    gap_level = to_numeric_series(df.get("gap_level", pd.Series(dtype=float)))
    # база для gap_size
    if mfe_base == "gap_level": base = gap_level
    elif mfe_base == "open":    base = open_
    elif mfe_base == "close":   base = close_
    else: base = gap_level
    gap_size_price = (base - open_).abs()
    atr_at_gap = atr14
    gap_size_atr = gap_size_price / atr_at_gap.replace(0, np.nan)

    # KM подготовка
    now_utc = pd.Timestamp.now(tz="UTC")
    closed = meta["closed"]
    durations_min = ((ft - ot).dt.total_seconds() / 60).where(closed)  # только закрытые для событий
    # для KM нужны и закрытые, и открытые → соберём единый durations и флаг события
    all_mask = gt.isin(["up","down"]) & ot.notna()
    dur_all = pd.Series(np.nan, index=df.index, dtype=float)
    ev_all  = pd.Series(False, index=df.index, dtype=bool)
    dur_all.loc[closed & all_mask] = durations_min.loc[closed & all_mask]
    dur_all.loc[(~closed) & all_mask] = ((now_utc - ot).dt.total_seconds() / 60).loc[(~closed) & all_mask]
    ev_all.loc[closed & all_mask] = True

    # сезонность
    hours = ot.dt.hour
    dows  = ot.dt.dayofweek  # 0=Mon

    out = dict(
        atr14=atr14, adx14=adx14,
        gap_size_price=gap_size_price, gap_size_atr=gap_size_atr,
        km_durations=dur_all, km_events=ev_all,
        hours=hours, dows=dows,
        closed_mask=closed, all_mask=all_mask
    )
    return out

def km_and_probs(meta_pack: dict, tfm: float):
    surv, med, p90 = km_survival(meta_pack["km_durations"], meta_pack["km_events"])
    # горизонты по барам и времени
    bars_h = [1,3,5,10,20]
    time_h_min = [60, 240, 1440]  # 1ч, 4ч, 1д
    out = {
        "KM — медиана времени до закрытия (мин)": med,
        "KM — P90 времени до закрытия (мин)": p90,
    }
    for b in bars_h:
        hmin = b * tfm
        out[f"P(fill ≤ {b} бар), %"] = round(km_prob_fill_leq(surv, hmin) * 100.0, 2) if surv else np.nan
    for tmin, label in zip(time_h_min, ["1 ч", "4 ч", "1 дн"]):
        out[f"P(fill ≤ {label}), %"] = round(km_prob_fill_leq(surv, tmin) * 100.0, 2) if surv else np.nan
    return out

def binned_probs_by_gap_atr(meta_pack: dict, tfm: float):
    gs = meta_pack["gap_size_atr"]
    # бины
    bins = [0, 0.5, 1, 1.5, 2, 9e9]
    labels = ["[0,0.5)","[0.5,1)","[1,1.5)","[1.5,2)","[2,+)"]
    cat = pd.cut(gs, bins=bins, labels=labels, include_lowest=True)
    d = pd.DataFrame(dict(cat=cat, t=meta_pack["km_durations"], e=meta_pack["km_events"]))
    out = {}
    for lab in labels:
        sub = d[d["cat"] == lab]
        if sub.empty:
            out[f"P(fill ≤ 10 бар | gap_ATR∈{lab}), %"] = np.nan
            out[f"KM — медиана (мин) | gap_ATR∈{lab}"] = np.nan
            continue
        surv, med, _ = km_survival(sub["t"], sub["e"])
        p10 = km_prob_fill_leq(surv, 10*tfm) if surv else np.nan
        out[f"P(fill ≤ 10 бар | gap_ATR∈{lab}), %"] = round(p10*100.0, 2) if pd.notna(p10) else np.nan
        out[f"KM — медиана (мин) | gap_ATR∈{lab}"] = med
    return out

def stopout_risk_and_expectations(mf: dict, atr_at_gap: pd.Series):
    out = {}
    # по закрытым
    for side in ["up","down","all"]:
        mae = mf[side]["mae_price"]
        mfe = mf[side]["mfe_price"]
        atr = atr_at_gap.reindex(mae.index)
        mae_atr = (mae / atr.replace(0, np.nan)).dropna()
        mfe_atr = (mfe / atr.replace(0, np.nan)).dropna()
        for k in [0.5,1.0,1.5]:
            if len(mae_atr):
                share = float((mae_atr > k).mean() * 100.0)
            else:
                share = np.nan
            if len(mae_atr):
                mask_ok = mae_atr <= k
                val = float((mfe_atr[mask_ok]).mean()) if mask_ok.any() else np.nan
            else:
                val = np.nan
            out[f"Стоп-риск k={k} ATR до закрытия {side}, %"] = round(share, 2) if pd.notna(share) else np.nan
            out[f"Ожидание MFE_ATR | не стопнут k={k} {side}"] = round(val, 4) if pd.notna(val) else np.nan
    return out

def post_fill_drift(df: pd.DataFrame, meta: dict, horizons_bars=(1,3,5,10)):
    """Возврат после закрытия: % и ATR-единицы через N баров"""
    ot = meta["ot"]; ft = meta["ft"]; tfm = meta["tfm"]
    close = to_numeric_series(df["close"])
    high = to_numeric_series(df["high"]); low = to_numeric_series(df["low"])
    atr14 = compute_atr(high, low, close, 14)

    # ряды по времени отсортированные
    idx_sorted = ot.sort_values().index
    ot_sorted = ot.loc[idx_sorted]
    close_sorted = close.loc[idx_sorted]
    atr_sorted = atr14.loc[idx_sorted]

    def find_val_at(ts):
        arr = ot_sorted.values
        pos = np.searchsorted(arr, ts.to_datetime64(), side="right") - 1
        if pos < 0: pos = 0
        return close_sorted.iloc[pos], atr_sorted.iloc[pos]

    out = {}
    for i in df.index:
        if not meta["closed"].loc[i]: continue
        fill_ts = ft.loc[i]
        base_close, base_atr = find_val_at(fill_ts)
        if pd.isna(base_close) or base_close == 0:
            continue
        for N in horizons_bars:
            end_ts = fill_ts + pd.Timedelta(minutes=N*tfm)
            end_close, end_atr = find_val_at(end_ts)
            if pd.isna(end_close): continue
            ret_pct = (end_close - base_close) / base_close * 100.0
            ret_atr = (end_close - base_close) / (base_atr if base_atr else np.nan)
            out.setdefault(f"Пост-закрытие +{N} бар — средний %,", []).append(ret_pct)
            out.setdefault(f"Пост-закрытие +{N} бар — P90 %,", []).append(ret_pct)
            out.setdefault(f"Пост-закрытие +{N} бар — доля >0 %,", []).append(1.0 if ret_pct>0 else 0.0)
            out.setdefault(f"Пост-закрытие +{N} бар — средний ATR,", []).append(ret_atr)

    agg = {}
    for k, vals in out.items():
        s = pd.Series(vals, dtype=float)
        if "P90" in k:
            agg[k.replace("P90", "P90")] = round(float(np.percentile(s, 90)), 4) if len(s) else np.nan
        elif "доля >0" in k:
            agg[k] = round(float(s.mean()*100.0), 2) if len(s) else np.nan
        elif "средний ATR" in k:
            agg[k] = round(float(s.mean()), 4) if len(s) else np.nan
        else:
            agg[k] = round(float(s.mean()), 4) if len(s) else np.nan
    return agg

def interarrival_stats(meta: dict):
    ot = meta["ot"]
    ok = ot.notna()
    if ok.sum() < 2:
        return {"Межгэповое время — медиана (мин)": np.nan, "Межгэповое время — P90 (мин)": np.nan}
    ts = ot[ok].sort_values()
    delta = (ts.values[1:] - ts.values[:-1]) / np.timedelta64(1, 'm')
    s = pd.Series(delta, dtype=float)
    return {
        "Межгэповое время — медиана (мин)": float(np.median(s)) if len(s) else np.nan,
        "Межгэповое время — P90 (мин)": float(np.percentile(s, 90)) if len(s) else np.nan
    }

def seasonality_block(meta_pack: dict, tfm: float):
    """Агрегаты: частота по DOW/часу и условная P(fill ≤ 10 бар) для самых выраженных корзин"""
    hours = meta_pack["hours"]; dows = meta_pack["dows"]
    d = pd.DataFrame({"h": hours, "dow": dows, "t": meta_pack["km_durations"], "e": meta_pack["km_events"]})
    # Hour
    hour_counts = d.groupby("h").size().sort_values(ascending=False)
    top_h = hour_counts.index[:1].tolist()
    out = {}
    for h in top_h:
        sub = d[d["h"] == h]
        surv, _, _ = km_survival(sub["t"], sub["e"])
        p10 = km_prob_fill_leq(surv, 10*tfm) if surv else np.nan
        out[f"Сезонность — Час {h}: частота, %"] = round(100.0 * len(sub) / len(d), 2) if len(d) else np.nan
        out[f"Сезонность — Час {h}: P(fill ≤ 10 бар), %"] = round(p10*100.0, 2) if pd.notna(p10) else np.nan
    # DOW
    dow_counts = d.groupby("dow").size().sort_values(ascending=False)
    if len(dow_counts):
        best_dow = int(dow_counts.index[0])
        sub = d[d["dow"] == best_dow]
        surv, _, _ = km_survival(sub["t"], sub["e"])
        p10 = km_prob_fill_leq(surv, 10*tfm) if surv else np.nan
        out[f"Сезонность — День {best_dow}: частота, %"] = round(100.0 * len(sub) / len(d), 2) if len(d) else np.nan
        out[f"Сезонность — День {best_dow}: P(fill ≤ 10 бар), %"] = round(p10*100.0, 2) if pd.notna(p10) else np.nan
    return out

def market_regime_block(meta_pack: dict):
    """Режим рынка (внутренний по инструменту): квантиль ATR и ADX → условные вероятности (заголовки под «Маркет-специфика/Режим рынка»)"""
    atr = meta_pack["atr14"]; adx = meta_pack["adx14"]
    d = pd.DataFrame({
        "atr": atr, "adx": adx,
        "t": meta_pack["km_durations"], "e": meta_pack["km_events"]
    }).dropna()
    out = {}
    if d.empty:
        out["Режим рынка — P(fill ≤ 10 бар) | ATR Q4, %"] = np.nan
        out["Режим рынка — P(fill ≤ 10 бар) | ADX Q4, %"] = np.nan
        return out
    # верхний квартиль
    atr_q4 = d["atr"].quantile(0.75)
    adx_q4 = d["adx"].quantile(0.75)
    for nm, thr in [("ATR", atr_q4), ("ADX", adx_q4)]:
        sub = d[d[nm.lower()] >= thr]
        if sub.empty:
            out[f"Режим рынка — P(fill ≤ 10 бар) | {nm} Q4, %"] = np.nan
            continue
        surv, _, _ = km_survival(sub["t"], sub["e"])
        p10 = km_prob_fill_leq(surv, 10 * meta_pack.get("tfm", 1.0))
        out[f"Режим рынка — P(fill ≤ 10 бар) | {nm} Q4, %"] = round(p10*100.0, 2) if pd.notna(p10) else np.nan
    return out

def funding_basis_block(df: pd.DataFrame, meta_pack: dict):
    """Для фьючерсов: условные вероятности по квантили funding_rate и basis (если есть колонки)."""
    funding = to_numeric_series(df.get("funding_rate", pd.Series(dtype=float)))
    spot = to_numeric_series(df.get("spot_price", pd.Series(dtype=float)))
    perp = to_numeric_series(df.get("close", pd.Series(dtype=float)))  # как прокси
    basis = perp - spot if spot.notna().sum() else pd.Series(np.nan, index=df.index)
    d = pd.DataFrame({
        "funding": funding,
        "basis": basis,
        "t": meta_pack["km_durations"],
        "e": meta_pack["km_events"]
    })
    out = {}
    for nm in ["funding","basis"]:
        s = d[nm]
        if s.notna().sum() < 10:
            out[f"Фьючерсы — P(fill ≤ 10 бар) | {nm} Q4, %"] = np.nan
            continue
        thr = s.quantile(0.75)
        sub = d[s >= thr]
        surv, _, _ = km_survival(sub["t"], sub["e"])
        p10 = km_prob_fill_leq(surv, 10 * meta_pack.get("tfm", 1.0)) if surv else np.nan
        out[f"Фьючерсы — P(fill ≤ 10 бар) | {nm} Q4, %"] = round(p10*100.0, 2) if pd.notna(p10) else np.nan
    return out

def orderflow_imbalance_block(df: pd.DataFrame, meta: dict):
    """Имбаланс в окне: taker_buy/(taker_buy+taker_sell). Порог 0.6 → условная вероятность."""
    tb = to_numeric_series(df.get("taker_base_vol",  pd.Series(dtype=float)))
    volume = to_numeric_series(df.get("volume", pd.Series(dtype=float)))
    # оценим buy/sell из tb и volume (приближённо): buy ≈ tb, sell ≈ volume - tb (если есть raw buy/sell — заменить)
    buy = tb
    sell = (volume - tb)
    denom = (buy + sell).replace(0, np.nan)
    imb = (buy / denom).clip(0, 1)
    d = pd.DataFrame({
        "imb": imb,
        "t": ((meta["ft"] - meta["ot"]).dt.total_seconds() / 60).where(meta["closed"]),  # закрытые
        "e": meta["closed"].astype(bool)
    })
    # Для KM на всем наборе (с учётом открытых)
    all_d = pd.DataFrame({"imb": imb, "t": meta["km_durations"], "e": meta["km_events"]}) if "km_durations" in meta else None
    out = {}
    if all_d is None or all_d["imb"].notna().sum() == 0:

        out["Окно — P(fill ≤ 10 бар) | imbalance>0.6, %"] = np.nan
        return out
    sub = all_d[all_d["imb"] > 0.6]
    if sub.empty:
        out["Окно — P(fill ≤ 10 бар) | imbalance>0.6, %"] = np.nan
        return out
    surv, _, _ = km_survival(sub["t"], sub["e"])
    tfm = meta.get("tfm", 1.0)
    p10 = km_prob_fill_leq(surv, 10*tfm) if surv else np.nan
    out["Окно — P(fill ≤ 10 бар) | imbalance>0.6, %"] = round(p10*100.0, 2) if pd.notna(p10) else np.nan
    return out

# =========================
# СБОРКА ОДНОЙ СТРОКИ
# =========================
def build_one_row(df: pd.DataFrame, symbol: str, interval: str, mfe_base: str,
                  prefer_existing: bool, now_equals_last: bool):
    meta = pick_minutes_bars(df, interval, prefer_existing=prefer_existing)
    minutes = meta["minutes"]; bars = meta["bars"]
    up_mask = meta["up_closed"]; dn_mask = meta["down_closed"]; all_mask = meta["closed"]

    gt = meta["gt"]; filled = meta["ft"].notna(); opened = ~filled
    closed_up = int((gt.eq("up") & filled).sum()); open_up = int((gt.eq("up") & opened).sum())
    closed_dn = int((gt.eq("down") & filled).sum()); open_dn = int((gt.eq("down") & opened).sum())
    count_up = closed_up + open_up; count_dn = closed_dn + open_dn; count_total = count_up + count_dn

    # --- агрегаты по закрытым (как раньше + % всего закрытых) ---
    total_closed = closed_up + closed_dn
    pct_closed_up = float(closed_up / total_closed * 100.0) if total_closed > 0 else np.nan
    pct_closed_dn = float(closed_dn / total_closed * 100.0) if total_closed > 0 else np.nan
    pct_closed_total = float(total_closed / count_total * 100.0) if count_total > 0 else np.nan

    def stat_bars(mask):
        s = pd.to_numeric(bars.where(mask), errors="coerce").dropna()
        if s.empty: return (np.nan, np.nan, np.nan)
        return (int(np.floor(s.min())), int(np.ceil(s.max())), round(float(s.mean()), 2))
    min_up_b, max_up_b, mean_up_b    = stat_bars(up_mask)
    min_dn_b, max_dn_b, mean_dn_b    = stat_bars(dn_mask)
    min_all_b, max_all_b, mean_all_b = stat_bars(all_mask)

    def t_pack(x):
        return stats_pack_num(x, want_pcts=True)
    t_up  = t_pack(minutes.where(up_mask).dropna())
    t_dn  = t_pack(minutes.where(dn_mask).dropna())
    t_all = t_pack(minutes.where(all_mask).dropna())

    tfm = meta["tfm"]
    def bars_qos(m, b):
        d = pd.concat([m, b], axis=1, keys=["m", "b"]).dropna()
        if d.empty: return (np.nan, np.nan, np.nan)
        real = d["m"]/tfm; rnd = real.round().astype(int)
        mae = float((real - d["b"]).abs().mean())
        exact = float((rnd == d["b"]).mean()*100); within1 = float(((rnd - d["b"]).abs()<=1).mean()*100)
        return (round(mae,3), round(exact,2), round(within1,2))
    q_up  = bars_qos(minutes.where(up_mask),  pd.to_numeric(bars.where(up_mask), errors="coerce"))
    q_dn  = bars_qos(minutes.where(dn_mask),  pd.to_numeric(bars.where(dn_mask), errors="coerce"))
    q_all = bars_qos(minutes.where(all_mask), pd.to_numeric(bars.where(all_mask), errors="coerce"))

    mf = aggregate_mfe_mae(df, interval, meta, mfe_base=mfe_base)

    open_agg = aggregate_open_ages(df, interval, meta, now_equals_last=now_equals_last)
    now_ts  = open_agg["now_utc"]
    last_ts = open_agg["last_ts"]

    from_now  = open_impulse_metrics(df, meta, mfe_base, end_ts=now_ts)
    from_last = open_impulse_metrics(df, meta, mfe_base, end_ts=last_ts)

    gapbar_stats = aggregate_gapbar_volume_delta_by_bucket(df, meta)
    oi_gapbar_stats = aggregate_gapbar_oi_by_bucket(df, meta)
    win_closed   = aggregate_window_volume_delta(df, meta, bucket="closed",    end_ts_global=None)
    win_open_now = aggregate_window_volume_delta(df, meta, bucket="open_now",  end_ts_global=now_ts)
    win_open_last= aggregate_window_volume_delta(df, meta, bucket="open_last", end_ts_global=last_ts)

    series_stats = compute_direction_run_stats(df)

    # ===== Новые блоки (13 пунктов) =====
    # Подготовка ATR/ADX, KM и вспомогательных наборов
    aux = build_gap_series(df, meta, mfe_base)
    aux["tfm"] = tfm
    # 1-2) KM и вероятности
    km_probs = km_and_probs(aux, tfm)
    # 3) ATR-нормировки MFE/MAE и gap_size
    atr_at_gap = aux["atr14"]
    gap_size_atr = aux["gap_size_atr"]
    # 4) Калибровка по gap_size_ATR
    calib = binned_probs_by_gap_atr(aux, tfm)
    # 5) Стоп-риски
    stoprisk = stopout_risk_and_expectations(mf, atr_at_gap)
    # 6) Пост-закрытие дрейф
    postdrift = post_fill_drift(df, meta, horizons_bars=(1,3,5,10))
    # 7) Интер-аррайвл
    interarr = interarrival_stats(meta)
    # 8) Сезонность
    season = seasonality_block(aux, tfm)
    # 9) Режим рынка (внутренний)
    regime = market_regime_block(aux)
    # 10) Фьючерсы funding/basis
    fundbasis = funding_basis_block(df, aux)
    # 11) Имбаланс окна
    imb = orderflow_imbalance_block(df, {**meta, **aux})

    # Сбор строки
    row = {
        "Монета": symbol, "ТФ": interval,

        "Гэпов вверх (всего)": count_up,
        " └─ Закрытых вверх": closed_up,
        " └─ Открытых вверх": open_up,
        "Гэпов вниз (всего)": count_dn,
        " └─ Закрытых вниз": closed_dn,
        " └─ Открытых вниз": open_dn,
        "Всего гэпов": count_total,

        # Итоги закрытых
        "Закрытые — всего, шт": total_closed,
        "Закрытые — всего, %": round(pct_closed_total, 2) if pd.notna(pct_closed_total) else np.nan,
        "Закрытые — вверх, %": round(pct_closed_up, 2) if pd.notna(pct_closed_up) else np.nan,
        "Закрытые — вниз, %":  round(pct_closed_dn, 2) if pd.notna(pct_closed_dn) else np.nan,

        # Бары до закрытия
        "Мин баров до закрытия ↑": min_up_b,
        "Макс баров до закрытия ↑": max_up_b,
        "Среднее баров до закрытия ↑": mean_up_b,
        "Мин баров до закрытия ↓": min_dn_b,
        "Макс баров до закрытия ↓": max_dn_b,
        "Среднее баров до закрытия ↓": mean_dn_b,
        "Мин баров до закрытия (все)": min_all_b,
        "Макс баров до закрытия (все)": max_all_b,
        "Среднее баров до закрытия (все)": mean_all_b,
    }

    def add_minutes_block(prefix, pack):
        hp = humanize_minutes_pack(pack)
        row[f"{prefix} — Мин"]     = hp["min"]
        row[f"{prefix} — Макс"]    = hp["max"]
        row[f"{prefix} — Среднее"] = hp["mean"]
        row[f"{prefix} — Медиана"] = hp["median"]
        row[f"{prefix} — P50"]     = hp["p50"]
        row[f"{prefix} — P90"]     = hp["p90"]
        row[f"{prefix} — P95"]     = hp["p95"]

    add_minutes_block("Минут до закрытия ↑", t_up)
    add_minutes_block("Минут до закрытия ↓", t_dn)
    add_minutes_block("Минут до закрытия (все)", t_all)

    mae_up, exact_up, within1_up = q_up
    mae_dn, exact_dn, within1_dn = q_dn
    mae_all, exact_all, within1_all = q_all
    row.update({
        "MAE баров ↑ (время→бары vs bars)": mae_up,
        "Точные совпадения ↑, %":           exact_up,
        "Совпадения ±1 бар ↑, %":           within1_up,
        "MAE баров ↓ (время→бары vs bars)": mae_dn,
        "Точные совпадения ↓, %":           exact_dn,
        "Совпадения ±1 бар ↓, %":           within1_dn,
        "MAE баров (все) (время→бары vs bars)": mae_all,
        "Точные совпадения (все), %":          exact_all,
        "Совпадения ±1 бар (все), %":          within1_all,
        "База для MFE/MAE": mf["base_label"],
    })

    # MFE/MAE в минутах → «человеческий» формат
    def mfe_mae_block_short(d, side):
        st_names = ["mfe_price", "mae_price", "mfe_pct", "mae_pct",
                    "mfe_min", "mae_min", "mfe_bars", "mae_bars"]
        labels = {
            "mfe_price": "mfe_price", "mae_price": "mae_price",
            "mfe_pct": "mfe_pct", "mae_pct": "mae_pct",
            "mfe_min": "time_to_mfe", "mae_min": "time_to_mae",
            "mfe_bars": "bars_to_mfe", "mae_bars": "bars_to_mae"
        }
        for nm in st_names:
            st = stats_pack_num(d[nm], want_pcts=True)
            base = labels[nm]
            if nm.endswith("_min"):
                hp = humanize_minutes_pack(st)
                row[f"{base}_mean_{side}"]   = hp["mean"]
                row[f"{base}_median_{side}"] = hp["median"]
                row[f"{base}_p90_{side}"]    = hp["p90"]
            else:
                row[f"{base}_mean_{side}"]   = round(st["mean"], 10) if pd.notna(st["mean"]) else np.nan
                row[f"{base}_median_{side}"] = round(st["median"], 10) if pd.notna(st["median"]) else np.nan
                row[f"{base}_p90_{side}"]    = round(st.get("p90", np.nan), 10) if pd.notna(st.get("p90", np.nan)) else np.nan

    mfe_mae_block_short(mf["up"], "up")
    mfe_mae_block_short(mf["down"], "down")
    mfe_mae_block_short(mf["all"], "all")

    def put_imp(prefix, data):
        for side, label in [("up","↑"), ("down","↓"), ("all","(все)")]:
            row[f"{prefix} — Импульс MFE/|MAE| среднее {label}"] = round(data[side]["mean"], 4) if pd.notna(data[side]["mean"]) else np.nan
            row[f"{prefix} — Импульс MFE/|MAE| медиана {label}"] = round(data[side]["median"], 4) if pd.notna(data[side]["median"]) else np.nan
            row[f"{prefix} — Импульс MFE/|MAE| P90 {label}"]    = round(data[side]["p90"], 4) if pd.notna(data[side]["p90"]) else np.nan
            row[f"{prefix} — Доля MFE > |MAE| {label}, %"]      = round(data[side]["share"], 2) if pd.notna(data[side]["share"]) else np.nan

    put_imp("Открытые (до СЕЙЧАС)", from_now)
    put_imp("Открытые (до ПОСЛЕДНЕЙ)", from_last)

    row.update({
        "Текущее время (UTC)": open_agg["now_utc"],
        "Последняя дата в таблице (UTC)": open_agg["last_ts"],
    })

    def flatten_open(prefix, bucket):
        row[f"{prefix} — Открытых ↑, шт"] = bucket["n_up"]
        row[f"{prefix} — Открытых ↓, шт"] = bucket["n_down"]
        row[f"{prefix} — Открытых всего, шт"] = bucket["n_all"]
        for side_label, key in [("↑", "up"), ("↓", "down"), ("(все)", "all")]:
            dd = bucket[key]
            row[f"{prefix} — Минут ср {side_label}"]   = dd["mean_min"]
            row[f"{prefix} — Минут мед {side_label}"]  = dd["median_min"]
            row[f"{prefix} — Минут макс {side_label}"] = dd["max_min"]
            row[f"{prefix} — P50 минут {side_label}"]  = dd.get("p50_min", "")
            row[f"{prefix} — P90 минут {side_label}"]  = dd.get("p90_min", "")
            row[f"{prefix} — P95 минут {side_label}"]  = dd.get("p95_min", "")
            row[f"{prefix} — Бары ср {side_label}"]    = dd["mean_bars"]
            row[f"{prefix} — Бары мед {side_label}"]   = dd["median_bars"]
            row[f"{prefix} — Бары макс {side_label}"]  = dd["max_bars"]
            row[f"{prefix} — Часы ср {side_label}"]    = dd["mean_hours"]
            row[f"{prefix} — Дни ср {side_label}"]     = dd["mean_days"]
            row[f"{prefix} — Месяцы ср {side_label}"]  = dd["mean_months"]
            row[f"{prefix} — Годы ср {side_label}"]    = dd["mean_years"]

    flatten_open("Открытые (до СЕЙЧАС)", open_agg["now"])
    flatten_open("Открытые (до ПОСЛЕДНЕЙ)", open_agg["last"])

    # Общие барные метрики
    row.update(gapbar_stats)
    # Фьючерсные OI-метрики
    row.update(oi_gapbar_stats)
    # Окно
    row.update(win_closed)
    row.update(win_open_now)
    row.update(win_open_last)
    # Серийность
    row.update(series_stats)

    # ====== Добавление НОВЫХ МЕТРИК в строку ======
    # 1-2
    row.update(km_probs)
    # 3
    row["gap_size_ATR — среднее"] = round(float(aux["gap_size_atr"].dropna().mean()), 4) if aux["gap_size_atr"].notna().any() else np.nan
    # 4
    row.update(calib)
    # 5
    row.update(stoprisk)
    # 6
    row.update(postdrift)
    # 7
    row.update(interarr)
    # 8
    row.update(season)
    # 9
    row.update(regime)
    # 10
    row.update(fundbasis)
    # 11
    row.update(imb)
    # 12 (мульти-ТФ) — заглушки
    row["Мульти-ТФ — флаг старший открытый гэп рядом"] = np.nan
    row["Мульти-ТФ — P(fill ≤ 10 бар) | конфликт направлений, %"] = np.nan

    return pd.DataFrame([row])

# =========================
# ГРУППЫ/ПОДГРУППЫ/ПОКАЗАТЕЛИ
# =========================
def make_grouped_columns_3(cols: list[str]) -> pd.MultiIndex:
    """
    Многоуровневые колонки: (Группа, Подгруппа, Показатель).
    Новые группы:
      - Вероятности закрытия (P(fill ≤ N)), KM-выживаемость
      - ATR-нормировка и калибровка по gap_size_ATR
      - Стоп-риски
      - После закрытия (дрейф)
      - Сезонность
      - Интер-аррайвл
      - Маркет-специфика → Режим рынка; Фьючерсы (OI); Фьючерсы (фандинг/бейсис)
      - Мульти-ТФ
      - Кросс-секционка (ранги) — добавляется после агрегации всех строк
    """
    def side_to_name(sig: str) -> str:
        return {"↑": "Вверх", "↓": "Вниз", "(все)": "Все"}.get(sig, sig)

    def detect_side(text: str) -> str | None:
        m = re.search(r"(↑|↓|\(все\))\s*$", text)
        return m.group(1) if m else None

    def strip_side(text: str) -> str:
        return re.sub(r"\s*(↑|↓|\(все\))\s*$", "", text).strip()

    out = []
    for c in cols:
        group = "Прочее"; subgroup = ""; label = c

        # базовые/время
        if c in ("Монета", "ТФ"):
            out.append(("Базовое", "", c)); continue
        if c in ("Текущее время (UTC)", "Последняя дата в таблице (UTC)"):
            out.append(("Время отчёта", "", c)); continue

        # Итоги закрытых
        if c.startswith("Закрытые — "):
            out.append(("Итоги закрытых", "", c)); continue

        # Вероятности / KM
        if c.startswith("P(fill") or c.startswith("KM — "):
            out.append(("Вероятности закрытия / KM", "", c)); continue

        # ATR-нормировка / калибровка
        if c.startswith("gap_size_ATR") or "gap_ATR∈" in c:
            out.append(("ATR-нормировка и калибровка", "", c)); continue

        # Стоп-риски
        if c.startswith("Стоп-риск") or c.startswith("Ожидание MFE_ATR"):
            out.append(("Стоп-риски", "", c)); continue

        # После закрытия
        if c.startswith("Пост-закрытие"):
            out.append(("После закрытия", "", c)); continue

        # Сезонность
        if c.startswith("Сезонность"):
            out.append(("Сезонность", "", c)); continue

        # Интер-аррайвл
        if c.startswith("Межгэповое время"):
            out.append(("Интер-аррайвл", "", c)); continue

        # Режим рынка (внутренний)
        if c.startswith("Режим рынка"):
            out.append(("Маркет-специфика", "Режим рынка", c)); continue

        # Фьючерсы: OI
        if "OI" in c:
            side_sig = detect_side(c) or "(все)"
            bucket_m = re.match(r"^\(([^)]+)\)", c)
            bucket_name = bucket_m.group(1) if bucket_m else "Все"
            out.append(("Маркет-специфика", f"Фьючерсы (OI) — {bucket_name} {side_to_name(side_sig)}", re.sub(r"^\([^)]+\)\s*", "", strip_side(c)))); continue

        # Фьючерсы: funding/basis
        if c.startswith("Фьючерсы — "):
            out.append(("Маркет-специфика", "Фьючерсы (фандинг/бейсис)", c)); continue

        # Имбаланс окна
        if c.startswith("Окно — P(fill"):
            out.append(("Окно гэпа", "Имбаланс", c)); continue

        # MFE/MAE/Качество/До закрытия/Окно гэпа (как было)
        if re.search(r"_(up|down|all)$", c):
            m = re.search(r"_(up|down|all)$", c); side = {"up": "Вверх", "down": "Вниз", "all": "Все"}[m.group(1)]
            out.append(("MFE/MAE (закрытые)", f"Закрытые {side}", re.sub(r"_(up|down|all)$","",c))); continue
        if c.startswith("MAE баров") or c.startswith("Точные совпадения") or "Совпадения ±1 бар" in c:
            side_sig = detect_side(c) or "(все)"
            out.append(("Качество (время→бары)", f"Закрытые {side_to_name(side_sig)}", strip_side(c))); continue
        if c.startswith("Мин баров до закрытия") or c.startswith("Минут до закрытия"):
            side_sig = detect_side(c) or "(все)"
            out.append(("До закрытия", f"Закрытые {side_to_name(side_sig)}", strip_side(c))); continue
        if c.startswith("(") and ("баре гэпа" in c or "барах гэпа" in c) and "OI" not in c:
            bucket = re.match(r"^\(([^)]+)\)", c).group(1) if re.match(r"^\(([^)]+)\)", c) else ""
            side_sig = detect_side(c) or "(все)"
            out.append(("Бар гэпа", f"{bucket} {side_to_name(side_sig)}", re.sub(r"^\([^)]+\)\s*", "", strip_side(c))))
            continue

        if c.startswith("(Закрытые ОКНО)") or c.startswith("(Открытые→СЕЙЧАС ОКНО)") or c.startswith("(Открытые→ПОСЛЕДНЯЯ ОКНО)"):
            bucket = c.split(")")[0].strip("(")
            side_sig = detect_side(c) or "(все)"
            out.append(("Окно гэпа", f"{bucket} {side_to_name(side_sig)}", re.sub(r"^\([^)]+\)\s*", "", strip_side(c))))
            continue

        # Серийность
        if c.startswith("Серии") or c.startswith("Распределение длин серий") or c.startswith("Вероятность смены") or c.startswith("Доля смен направления"):
            out.append(("Серийность", "", c)); continue

        # Открытые (импульсы) — уже помечены префиксом
        if c.startswith("Открытые (до СЕЙЧАС)") or c.startswith("Открытые (до ПОСЛЕДНЕЙ)"):
            side_sig = detect_side(c) or "(все)"
            prefix = "СЕЙЧАС" if "СЕЙЧАС" in c else "ПОСЛЕДНЕЙ"
            out.append(("Открытые — импульс", prefix + " " + side_to_name(side_sig), strip_side(c))); continue

        # Открытые — возраст
        if c.startswith("Открытые (до СЕЙЧАС) — ") or c.startswith("Открытые (до ПОСЛЕДНЕЙ) — "):
            prefix = "СЕЙЧАС" if "СЕЙЧАС" in c else "ПОСЛЕДНЕЙ"
            # выделим подгруппу по стрелке (если есть)
            side_sig = detect_side(c)
            sublabel = strip_side(re.sub(r"^Открытые \(до [^)]+\) — ", "", c))
            if side_sig:
                out.append(("Открытые — возраст", prefix + " " + side_to_name(side_sig), sublabel))
            else:
                out.append(("Открытые — возраст", prefix, sublabel))
            continue

        # По умолчанию
        out.append(("Прочее", "", c))

    return pd.MultiIndex.from_tuples(out, names=["Группа", "Подгруппа", "Показатель"])


# =========================
# ЭКСПОРТ В EXCEL С ГРУППАМИ/ПОДГРУППАМИ И ЖИРНЫМИ ЛИНИЯМИ
# =========================
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def dataframe_to_multiindex(df: pd.DataFrame) -> pd.DataFrame:
    """Преобразует плоские колонки df в MultiIndex (Группа, Подгруппа, Показатель), ничего не удаляя."""
    cols = list(df.columns)
    mi = make_grouped_columns_3(cols)
    out = df.copy()
    out.columns = mi
    return out

def style_header_and_borders(ws, n_header_rows: int, n_data_rows: int, n_cols: int):
    """Стилизует шапку (жирный, центр), и рисует вертикальные границы между ГРУППАМИ/ПОДГРУППАМИ.
       n_header_rows — число строк в заголовке (уровней MultiIndex), обычно 3.
       Первая колонка — индекс (A), начинаем стили с B.
    """
    thin = Side(border_style="thin", color="808080")
    med  = Side(border_style="medium", color="404040")
    thick= Side(border_style="thick", color="000000")

    # Шапка: жирный, заливка
    header_fill_group = PatternFill("solid", fgColor="E6F2FF")
    header_fill_sub   = PatternFill("solid", fgColor="F5FAFF")

    for r in range(1, n_header_rows+1):
        for c in range(2, n_cols+2):  # +1 из-за индекса, +1 т.к. включительно
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # Заливка: верхний уровень темнее
            cell.fill = header_fill_group if r == 1 else header_fill_sub

    # Вертикальные границы: считываем фактические тексты шапки
    # Уровень 1 (группы) — строка 1, уровень 2 (подгруппы) — строка 2
    def col_text(row, col):
        return ws.cell(row=row, column=col).value

    # Собираем изменения групп/подгрупп, чтобы рисовать границы слева
    def add_left_border(col_idx, style_side):
        # от шапки до конца данных
        for r in range(1, n_header_rows + n_data_rows + 1):
            cell = ws.cell(row=r, column=col_idx)
            b = cell.border
            cell.border = Border(left=style_side, right=b.right, top=b.top, bottom=b.bottom)

    # Границы для групп (строка 1)
    prev_group = None
    for c in range(2, n_cols+2):
        g = col_text(1, c)
        if g != prev_group:
            add_left_border(c, thick)  # толстая линия в начале новой группы
            prev_group = g
    # Правый край таблицы
    add_left_border(n_cols+2, thick)  # визуально как правая граница

    # Границы для подгрупп (строка 2) — средняя толщина
    prev_sub = None
    for c in range(2, n_cols+2):
        s = col_text(2, c)
        if s != prev_sub:
            add_left_border(c, med)
            prev_sub = s

    # Горизонтальные границы снизу шапки
    for c in range(2, n_cols+2):
        cell = ws.cell(row=n_header_rows, column=c)
        b = cell.border
        cell.border = Border(top=b.top, left=b.left, right=b.right, bottom=thick)

    # Выравнивания данных: перенос по словам
    for r in range(n_header_rows+1, n_header_rows + n_data_rows + 1):
        for c in range(2, n_cols+2):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    # Заморозка панелей: ниже шапки и после индекса
    ws.freeze_panes = ws.cell(row=n_header_rows+1, column=2)

def autosize_columns(ws, max_width=60):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            v = str(val)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[letter].width = min(max_len * 0.95 + 2, max_width)

def export_with_styles(df_multi: pd.DataFrame, out_xlsx: str, sheet_name="Итог_1_строка"):
    # ВАЖНО: для MultiIndex pandas требует index=True (иначе NotImplementedError)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as wr:
        df_multi.to_excel(wr, sheet_name=sheet_name, index=True)
        ws = wr.sheets[sheet_name]

        # Стайлинг
        n_header_rows = len(df_multi.columns.levels)
        n_data_rows   = df_multi.shape[0]
        n_cols        = df_multi.shape[1]

        # Спрячем индекс визуально (сузим)
        ws.column_dimensions[get_column_letter(1)].width = 3

        style_header_and_borders(ws, n_header_rows, n_data_rows, n_cols)
        autosize_columns(ws, max_width=48)


# =========================
# ОБРАБОТКА ФАЙЛА (как было)
# =========================
def process_one_file(file_path: str, mfe_base: str, prefer_existing: bool, now_equals_last: bool) -> pd.DataFrame:
    try:
        fp = smart_resolve_path(file_path)
        df = pd.read_excel(fp)
        for col in ["gap_type", "open_time", "gap_filled_time", "high", "low", "open", "close", "volume"]:
            if col not in df.columns:
                raise KeyError(f"Отсутствует обязательная колонка: {col}")
        symbol, interval = extract_symbol_interval(fp)
        one = build_one_row(df, symbol, interval, mfe_base, prefer_existing, now_equals_last)
        return one
    except Exception as e:
        print(f"✖ Ошибка в файле: {file_path}\n{e}\n{traceback.format_exc()}")
        try:
            sym, itv = extract_symbol_interval(file_path)
        except Exception:
            sym, itv = "UNKNOWN", "UNKNOWN"
        return pd.DataFrame([{
            "Монета": sym, "ТФ": itv, "Всего гэпов": np.nan, "ERROR": str(e)
        }])

# =========================
# ДОБАВЛЕНО: батч-обработка по (exchange, market)
# =========================
def collect_files_for_root(input_root: str) -> list[str]:
    """
    Возвращает список файлов full_with_gaps.xlsx в структуре:
    <root>\exchange=<ex>\market=<mk>\symbol=*\interval=*\full_with_gaps.xlsx
    Если input_root уже внутри exchange=/market= — будет работать как раньше.
    """
    input_root = os.path.normpath(input_root)
    # если путь уже указывает на конкретную биржу/рынок
    if ("exchange=" in input_root) and ("market=" in input_root):
        pattern = os.path.join(input_root, "symbol=*", "interval=*", "full_with_gaps.xlsx")
        return sorted(glob(pattern))
    # иначе собираем для всех exchange=* / market=*
    pattern = os.path.join(input_root, "exchange=*", "market=*", "symbol=*", "interval=*", "full_with_gaps.xlsx")
    return sorted(glob(pattern))

def run_batch_for_dataset(files: list[str],
                          mfe_base: str,
                          prefer_existing: bool,
                          now_equals_last: bool,
                          output_dir: str,
                          workers: int,
                          dry_run: bool):
    """
    Обрабатывает список файлов ОДНОГО датасета (одной биржи и одного рынка) и сохраняет XLSX
    с именем GepAnalitic-Dataset-Core_<exchange>_<market>_<ts>.xlsx
    """
    if not files:
        return

    # определим exchange/market по первому файлу
    ex, mk = extract_exchange_market(files[0])

    print(f"\n=== Обработка датасета: exchange={ex}, market={mk} ===")
    print(f"Файлов: {len(files)}")
    print(f"Пример: {files[0]}")

    rows = [] 
    if workers and workers > 1:
        print(f"Запуск в параллель {workers} процессов…")
        with ProcessPoolExecutor(max_workers=workers) as ex_pool:
            fut2file = {
                ex_pool.submit(process_one_file, f, mfe_base, prefer_existing, now_equals_last): f
                for f in files
            }
            for i, fut in enumerate(as_completed(fut2file), 1):
                f = fut2file[f]
                try:
                    df_row = fut.result()
                    rows.append(df_row)
                except Exception as e:
                    print(f"✖ Критическая ошибка в задаче для {f}: {e}")
                if i % 50 == 0:
                    print(f"…обработано {i}/{len(files)}")
    else:
        print("Последовательная обработка…")
        for i, f in enumerate(files, 1):
            df_row = process_one_file(f, mfe_base, prefer_existing, now_equals_last)
            rows.append(df_row)
            if i % 50 == 0:
                print(f"…обработано {i}/{len(files)}")

    if not rows:
        print("⚠ Нет результатов для сохранения (пропускаю датасет).")
        return

    result = pd.concat(rows, ignore_index=True)
    # Базовые колонки впереди
    col_order = ["Монета", "ТФ"]
    other_cols = [c for c in result.columns if c not in col_order]
    result = result[col_order + other_cols]

    print("\nПревью первых 10 строк итоговой таблицы:\n")
    with pd.option_context('display.max_columns', 20, 'display.width', 200):
        print(result.head(10).to_string(index=False))

    if dry_run:
        print("\n(DRY-RUN) Файл не сохранён.")
        return

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_xlsx = os.path.join(output_dir, f"GepAnalitic-Dataset-Core_{ex}_{mk}_{ts}.xlsx")

    # Конвертируем в MultiIndex и аккуратно сохраняем (index=True)
    multi_df = dataframe_to_multiindex(drop_tz_for_excel(result))
    export_with_styles(multi_df, out_xlsx, sheet_name="Итог_1_строка")

    print(f"\n✔ Итог сохранён: {out_xlsx}")
    print(f"Всего строк: {len(result)}")

# =========================
# MAIN: батч + экспорт в Excel (расширено под spot/другие биржи)
# =========================
def main():
    parser = argparse.ArgumentParser(description="Батч: одна сводная строка на файл full_with_gaps.xlsx. Итог в Excel.")
    parser.add_argument("--input-root", default=DEFAULT_INPUT_ROOT,
                        help="Корень данных. Может быть как конкретный exchange=..\\market=.., так и общий корень gap-analytics-xlsx-full.")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR,
                        help="Куда сохранить общий Excel")
    parser.add_argument("--mfe-base", choices=["gap_level", "open", "close"], default="gap_level",
                        help="База для MFE/MAE (по умолчанию gap_level)")
    parser.add_argument("--prefer-existing", dest="prefer_existing", action="store_true", default=True,
                        help="Использовать готовые gap_time_to_fill/gap_bars_to_fill (если есть)")
    parser.add_argument("--recompute", dest="prefer_existing", action="store_false",
                        help="Игнорировать готовые поля и пересчитать из времени")
    parser.add_argument("--now-equals-last", action="store_true",
                        help="Считать 'сейчас' равным последней дате в таблице (детерминирует отчёт)")
    parser.add_argument("--workers", type=int, default=0,
                        help="Число процессов (0 или 1 — без параллели)")
    parser.add_argument("--max-files", type=int, default=0,
                        help="Ограничить количество обрабатываемых файлов (для теста). 0 — без лимита.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Не сохранять Excel, просто показать превью/кол-во.")

    # >>> ДОБАВЛЕНО: фильтры по биржам/рынкам при общем корне
    parser.add_argument("--exchanges", type=str, default="",
                        help="Список бирж через запятую (например: binance,bybit,okx). Пусто — все, что найдётся.")
    parser.add_argument("--markets", type=str, default="",
                        help="Список рынков через запятую (например: futures,spot). Пусто — все, что найдётся.")

    args = parser.parse_args()

    input_root = os.path.normpath(args.input_root)
    output_dir = os.path.normpath(args.output_dir)
    os.makedirs(output_dir, exist_ok=True)

    # Если задан конкретный exchange=..\\market=.. в input_root — поведение как раньше (ОДИН датасет)
    if ("exchange=" in input_root) and ("market=" in input_root):
        pattern = os.path.join(input_root, "symbol=*", "interval=*", "full_with_gaps.xlsx")
        files = sorted(glob(pattern))
        if args.max_files and args.max_files > 0:
            files = files[:args.max_files]
        if not files:
            raise FileNotFoundError(f"Не найдено файлов по шаблону: {pattern}")
        run_batch_for_dataset(
            files=files,
            mfe_base=args.mfe_base,
            prefer_existing=args.prefer_existing,
            now_equals_last=args.now_equals_last,
            output_dir=output_dir,
            workers=args.workers,
            dry_run=args.dry_run
        )
        return

    # Иначе: общий корень → собираем все датасеты и при необходимости фильтруем
    all_files = collect_files_for_root(input_root)
    if not all_files:
        raise FileNotFoundError(f"Не найдено файлов в корне: {input_root}")

    # Группировка по (exchange, market)
    from collections import defaultdict
    buckets: dict[tuple[str,str], list[str]] = defaultdict(list)
    for f in all_files:
        ex, mk = extract_exchange_market(f)
        buckets[(ex, mk)].append(f)

    # Фильтрация по --exchanges/--markets (если заданы)
    sel_ex = set([x.strip().lower() for x in args.exchanges.split(",") if x.strip()]) if args.exchanges else None
    sel_mk = set([x.strip().lower() for x in args.markets.split(",") if x.strip()]) if args.markets else None

    # Пробегаем по каждому датасету и сохраняем отдельный XLSX
    total_sets = 0
    for (ex, mk), files in sorted(buckets.items()):
        if sel_ex and ex.lower() not in sel_ex:
            continue
        if sel_mk and mk.lower() not in sel_mk:
            continue
        if args.max_files and args.max_files > 0:
            files = files[:args.max_files]

        run_batch_for_dataset(
            files=files,
            mfe_base=args.mfe_base,
            prefer_existing=args.prefer_existing,
            now_equals_last=args.now_equals_last,
            output_dir=output_dir,
            workers=args.workers,
            dry_run=args.dry_run
        )
        total_sets += 1

    if total_sets == 0:
        print("⚠ После фильтрации нечего обрабатывать (проверьте --exchanges / --markets).")


if __name__ == "__main__":
    main()



